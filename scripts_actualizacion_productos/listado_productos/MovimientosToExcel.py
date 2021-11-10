import xlwt

import xlsxwriter,xlrd
import sys,os.path
#import xlsxwriter

import xlrd
from xlrd import open_workbook

import xlutils

from xlutils import  copy

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


#esta clase tiene 2 atributos
#wb que guarda el arcivo
#fila que fuarda el numero de fila en la que va
class MovimientosToExcel:




    #crea el archivo ecxel y se asigna al atributo wb del objeto MovimientosToExcel
	def __init__(self,name):


		if  name == "MAHINDRA":
			self.wb = xlwt.Workbook()
			self.ws = self.wb.add_sheet(name, cell_overwrite_ok=True)

		elif name == "ALIANZAS":
			self.wb = xlrd.open_workbook("listadoDeProductos.xls")
			print(self.wb.sheet_names())
			self.wss = self.wb.sheet_by_name("MAHINDRA")
			print(self.wss.name)

			self.wb = xlwt.Workbook()
			self.ws = self.wb.add_sheet("MAHINDRA", cell_overwrite_ok=True)

			for i in range(self.wss.nrows):

				for j in range(self.wss.ncols):
					cell_value = self.wss.cell_value(i, j, )  # c.. Obtener el valor de la columna j en la fila i
					self.ws.write(i, j, cell_value)  # Escriba el valor obtenido en el rango correspondiente del archivo

			self.ws = self.wb.add_sheet(name, cell_overwrite_ok=True)

		elif name == "ZEUS":

			#recupera objeto worbook con hoja llamada mahindra
			self.archivoanterior = xlrd.open_workbook("listadoDeProductos.xls")
			#print(self.wb.sheet_names())
			self.wsma = self.archivoanterior.sheet_by_name("MAHINDRA")
			#print(self.wss.name)

            #nuevo libro donde se giardar aen primera isntancia lo que se recupere del anterior archivo y se ageguen los listados d eproductos d la 3bd
			self.wb = xlwt.Workbook()
			self.ws = self.wb.add_sheet("MAHINDRA", cell_overwrite_ok=True)

            #se recorre el antiguo archivo y se le asigna el valor de la hoja mahindra al nuevo archivo
			#mahindra
			for i in range(self.wsma.nrows):

				for j in range(self.wsma.ncols):
					cell_value = self.wsma.cell_value(i, j, )  # c.. Obtener el valor de la columna j en la fila i
					self.ws.write(i, j, cell_value)  # Escriba el valor obtenido en el rango correspondiente del archivo
				# format.set_align('vcenter')  # Establecer la alineación vertical de la celda

			print("ya paso mahindra")


            #ALIANZAS
			# crea una nueva hoja llamada alianzas en el archivo nuevo
			self.ws = self.wb.add_sheet("ALIANZAS", cell_overwrite_ok=True)

			# se recorre el antiguo archivo y se le asigna el valor de la hoja alianzas al nuevo archivo
			self.wsalianzas = self.archivoanterior.sheet_by_name("ALIANZAS")

			for i in range(self.wsalianzas.nrows):

				for j in range(self.wsalianzas.ncols):
					cell_value = self.wsalianzas.cell_value(i, j, )  # c.. Obtener el valor de la columna j en la fila i
					self.ws.write(i, j, cell_value)  # Escriba el valor obtenido en el rango correspondiente del archivo


			self.ws = self.wb.add_sheet("ZEUS", cell_overwrite_ok=True)



		#se crear un array que contiene la cabezera d cada columna

		columnas = ["OPERADOR","TIPO DE PRODUCTO","ID MOVIIRED","DESCRIPCION","VALOR","EAN","ID OPERADOR"]

		print("pasamos")




        #se recorre el array columnas y se agrega en cada columna de la fila 1
		c = 0
		for columna in columnas:
			self.ws.write(0, c, columna)
			c = c + 1
		'''for columna in columnas:

			self.ws[f'[B]{C}']=COLUMNA
			c = c + 1'''


        #atributo del objeto ovimientosToExcel que indica que liena del archivo excel se esta iterando
		self.fila = 1

    #imprime cada atributo del objeto movimiento en cad acolumna de una fila del archivo excel
	def agregarItem(self,item):

		self.ws.write(self.fila, 0, item.operator_name)
		self.ws.write(self.fila, 1, item.tipodeproducto)
		self.ws.write(self.fila, 2, item.idmoviired)
		self.ws.write(self.fila, 3, item.descripcion)
		self.ws.write(self.fila, 4, item.valor)
		self.ws.write(self.fila, 5, item.ean)
		self.ws.write(self.fila, 6, item.idoperador)

		self.fila = self.fila + 1






	#en el parametro se le envia el  nombre del archivo con e l que se va guardar
	def guardarPlanilla(self,archivo):
		self.wb.save(archivo)
		print ("Generado")

